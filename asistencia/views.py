from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from .forms import RegistroAsistenciaForm
from .models import RegistroAsistencia


def formulario_asistencia(request):
    if request.method == "POST":
        form = RegistroAsistenciaForm(request.POST)
        if form.is_valid():
            asistencia = form.save(commit=False)
            asistencia.fecha = timezone.localdate()
            asistencia.save()
            messages.success(request, "asistencia registrada correctametne")
            return redirect("asistencia:confirmacion")
    else:
        form = RegistroAsistenciaForm()
    return render(request, "asistencia/formulario.html", {"form": form})


def confirmacion_asistencia(request):
    return render(request, "asistencia/confirmacion.html")


class ExportarExcelPorFechaView(View):
    """GET /excel/AAAA-MM-DD/ — descarga .xlsx con alumnos de esa fecha."""

    def get(self, request, fecha_str):
        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Fecha inválida. Use el formato AAAA-MM-DD.", status=400)

        registros = RegistroAsistencia.objects.filter(fecha=fecha).order_by(
            "apellido", "nombre"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        titulo = f"Asistencias correspondientes al dia {fecha.isoformat()}"
        ws.merge_cells("A1:D1")
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["DNI", "Apellido", "Nombre"]
        ws.append([])
        ws.append(headers)
        for cell in ws[3]:
            cell.font = Font(bold=True)

        vistos = set()
        for r in registros:
            clave = (r.dni, r.apellido, r.nombre, r.fecha)
            if clave in vistos:
                continue
            vistos.add(clave)
            ws.append(
                [
                    r.dni,
                    r.apellido,
                    r.nombre,
                ]
            )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        fname = f"asistencia_{fecha_str}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        return response




